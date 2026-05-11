#!/usr/bin/env python3
"""
One-shot importer: flatten the original outreach workbooks into
docs/outreach/contacts.csv (the SSOT going forward).

After this CSV becomes canonical the script is kept for provenance only.
Re-run if a fresh import from the source workbooks is needed.

Usage:
    python3 scripts/outreach/import_xlsx.py [path-to-xlsx ...]

If no paths are passed, both default workbooks are imported in order:
    ~/Desktop/Layout_Outreach_Updated.xlsx          (broad list, ~283 rows)
    ~/Downloads/layout-design-influencer-outreach.xlsx  (Figma educators, 39 rows)

Output path: docs/outreach/contacts.csv (relative to repo root).

Dedupe is by handle (case-insensitive), falling back to (sheet, name slug).
First file in wins on duplicates so re-importing the broad list first preserves
the existing IDs.
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    sys.exit(1)


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX_PATHS = [
    Path.home() / "Desktop" / "Layout_Outreach_Updated.xlsx",
    Path.home() / "Downloads" / "layout-design-influencer-outreach.xlsx",
]
SEEDS_DIR = REPO_ROOT / "docs" / "outreach" / "seeds"
OUTPUT = REPO_ROOT / "docs" / "outreach" / "contacts.csv"

FIELDS = [
    "id",
    "name",
    "handle",
    "platform",
    "followers",
    "followers_raw",
    "tier",
    "audience_category",
    "persona",
    "priority",
    "partner_tier",  # standard | flagship — drives commission + DM stance
    "status",
    "outreach_channel",
    "outreach_date",
    "batch_code",
    "response",
    "notes",
    "source_sheet",
    "source_workbook",
]


def parse_followers(raw: Any) -> tuple[int | None, str]:
    """Return (numeric followers, raw string) from a cell like '1.8M', '~50K', '40,297', None."""
    if raw is None:
        return None, ""
    s = str(raw).strip()
    if not s or s in {"-", "—"}:
        return None, s
    cleaned = s.replace(",", "").replace("~", "").replace("+", "").strip()
    m = re.match(r"([\d.]+)\s*([KkMm])?", cleaned)
    if not m:
        return None, s
    num = float(m.group(1))
    unit = (m.group(2) or "").lower()
    if unit == "k":
        num *= 1_000
    elif unit == "m":
        num *= 1_000_000
    return int(num), s


def tier_from_followers(n: int | None) -> str:
    if n is None:
        return ""
    if n < 10_000:
        return "nano"
    if n < 100_000:
        return "micro"
    if n < 500_000:
        return "macro"
    return "mega"


def normalise_handle(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s or s in {"-", "—"}:
        return ""
    if s.startswith("http"):
        return s
    if not s.startswith("@") and " " not in s and "/" not in s:
        return "@" + s.lstrip("@")
    return s


def normalise_priority(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().upper()
    if s in {"HIGH", "P0", "T1", "TIER 1", "1"}:
        return "P0"
    if s in {"MEDIUM", "MED", "P1", "T2", "TIER 2", "2"}:
        return "P1"
    if s in {"LOW", "P2", "T3", "TIER 3", "3"}:
        return "P2"
    if s in {"P3", "T4", "TIER 4", "4"}:
        return "P3"
    return ""


def normalise_tier_label(raw: Any) -> str:
    """For workbooks that provide an explicit tier label like 'Tier 1'."""
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    if "tier 1" in s or s == "1":
        return "tier-1"
    if "tier 2" in s or s == "2":
        return "tier-2"
    if "tier 3" in s or s == "3":
        return "tier-3"
    if "tier 4" in s or s == "4":
        return "tier-4"
    return ""


def cell(row: tuple, idx: int) -> Any:
    return row[idx] if idx < len(row) else None


def first_nonempty(row: tuple, *indices: int) -> str:
    for i in indices:
        v = cell(row, i)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def row_is_empty(row: tuple) -> bool:
    return not any(c is not None and str(c).strip() for c in row)


def category_for_sheet(sheet: str) -> str:
    return {
        "Twitter-X": "design-creator",
        "Top Designers": "design-creator",
        "YouTube-New": "design-creator",
        "Youtube": "design-creator",
        "TikTok-Instagram": "design-creator",
        "Figma Team": "figma-team",
        "Podcasts": "podcast",
        "Newsletters": "newsletter",
        "Discord Servers": "discord",
        "Paid Promotions": "paid-promotion",
        "Investors": "investor",
        # Second workbook
        "Outreach List": "design-creator",
    }.get(sheet, "")


def slug(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "contact"


def process_workbook(
    xlsx_path: Path,
    rows: list[dict[str, Any]],
    seen_handles: set[str],
    seen_names: set[str],
) -> int:
    """Append rows from one workbook. Returns count of duplicates skipped."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    skipped_dupes = 0
    workbook_tag = xlsx_path.stem

    for sheet_name in wb.sheetnames:
        if sheet_name == "Dashboard":
            continue
        ws = wb[sheet_name]
        category = category_for_sheet(sheet_name)

        for r in ws.iter_rows(min_row=2, values_only=True):
            if row_is_empty(r):
                continue

            name = ""
            handle = ""
            platform = ""
            followers_raw_cell: Any = None
            priority_raw: Any = None
            notes_parts: list[str] = []

            if sheet_name == "Twitter-X":
                priority_raw = cell(r, 0)
                name = first_nonempty(r, 1)
                handle = normalise_handle(cell(r, 2))
                platform = "Twitter/X"
                followers_raw_cell = cell(r, 3)
                if cell(r, 4):
                    notes_parts.append(f"category: {cell(r, 4)}")
                if cell(r, 6):
                    notes_parts.append(f"angle: {cell(r, 6)}")
            elif sheet_name == "Investors":
                priority_raw = cell(r, 0)
                name = first_nonempty(r, 1)
                handle = normalise_handle(cell(r, 3))
                platform = "Twitter/X"
                followers_raw_cell = cell(r, 4)
                if cell(r, 2):
                    notes_parts.append(f"type: {cell(r, 2)}")
                if cell(r, 5):
                    notes_parts.append(f"paper-link: {cell(r, 5)}")
                if cell(r, 6):
                    notes_parts.append(f"angle: {cell(r, 6)}")
            elif sheet_name == "TikTok-Instagram":
                priority_raw = cell(r, 0)
                name = first_nonempty(r, 1)
                platform = str(cell(r, 2) or "").strip()
                handle = normalise_handle(cell(r, 3))
                followers_raw_cell = cell(r, 4)
                if cell(r, 5):
                    notes_parts.append(f"category: {cell(r, 5)}")
            elif sheet_name == "YouTube-New":
                priority_raw = cell(r, 0)
                name = first_nonempty(r, 1)
                handle = normalise_handle(cell(r, 6))  # Twitter handle column
                platform = "YouTube"
                followers_raw_cell = cell(r, 3)
                if cell(r, 2):
                    notes_parts.append(f"youtube: {cell(r, 2)}")
                if cell(r, 4):
                    notes_parts.append(f"category: {cell(r, 4)}")
                if cell(r, 5):
                    notes_parts.append(f"angle: {cell(r, 5)}")
            elif sheet_name == "Youtube":
                name = first_nonempty(r, 0)
                if not name:
                    continue
                handle = normalise_handle(cell(r, 4))  # Twitter
                platform = "YouTube"
                followers_raw_cell = cell(r, 3)
                if cell(r, 2):
                    notes_parts.append(f"youtube: {cell(r, 2)}")
                if cell(r, 1):
                    notes_parts.append(f"email: {cell(r, 1)}")
            elif sheet_name == "Figma Team":
                priority_raw = cell(r, 0)
                name = first_nonempty(r, 1)
                handle = normalise_handle(cell(r, 3))
                platform = "Twitter/X (Figma)"
                followers_raw_cell = cell(r, 4)
                if cell(r, 2):
                    notes_parts.append(f"role: {cell(r, 2)}")
                if cell(r, 6):
                    notes_parts.append(f"angle: {cell(r, 6)}")
            elif sheet_name == "Top Designers":
                priority_raw = cell(r, 0)
                name = first_nonempty(r, 1)
                platform = str(cell(r, 2) or "").strip()
                handle = normalise_handle(cell(r, 3))
                followers_raw_cell = cell(r, 4)
                if cell(r, 5):
                    notes_parts.append(f"role: {cell(r, 5)}")
                if cell(r, 6):
                    notes_parts.append(f"why: {cell(r, 6)}")
            elif sheet_name == "Podcasts":
                priority_raw = cell(r, 0)
                name = first_nonempty(r, 1)
                platform = "Podcast"
                if cell(r, 2):
                    notes_parts.append(f"hosts: {cell(r, 2)}")
                followers_raw_cell = cell(r, 3)
                if cell(r, 4):
                    notes_parts.append(f"category: {cell(r, 4)}")
                if cell(r, 6):
                    notes_parts.append(f"model: {cell(r, 6)}")
            elif sheet_name == "Newsletters":
                priority_raw = cell(r, 0)
                name = first_nonempty(r, 1)
                platform = "Newsletter"
                if cell(r, 2):
                    notes_parts.append(f"publisher: {cell(r, 2)}")
                followers_raw_cell = cell(r, 3)
                if cell(r, 4):
                    notes_parts.append(f"rate: {cell(r, 4)}")
                if cell(r, 5):
                    notes_parts.append(f"category: {cell(r, 5)}")
            elif sheet_name == "Paid Promotions":
                priority_raw = cell(r, 0)
                name = first_nonempty(r, 1)
                platform = str(cell(r, 2) or "").strip()
                followers_raw_cell = cell(r, 4)
                if cell(r, 3):
                    notes_parts.append(f"handle: {cell(r, 3)}")
                if cell(r, 5):
                    notes_parts.append(f"rate: {cell(r, 5)}")
                if cell(r, 6):
                    notes_parts.append(f"format: {cell(r, 6)}")
                if cell(r, 7):
                    notes_parts.append(f"contact: {cell(r, 7)}")
            elif sheet_name == "Discord Servers":
                priority_raw = cell(r, 0)
                name = first_nonempty(r, 1)
                platform = "Discord"
                followers_raw_cell = cell(r, 2)
                if cell(r, 3):
                    notes_parts.append(f"invite: {cell(r, 3)}")
                if cell(r, 4):
                    notes_parts.append(f"category: {cell(r, 4)}")
                if cell(r, 5):
                    notes_parts.append(f"why: {cell(r, 5)}")
                if cell(r, 6):
                    notes_parts.append(f"how: {cell(r, 6)}")
            elif sheet_name == "Outreach List":
                # Second workbook: Figma educators / design creators
                priority_raw = cell(r, 0)
                handle_or_name_raw = first_nonempty(r, 1)
                real_name = first_nonempty(r, 2)
                name = real_name or handle_or_name_raw
                # Extract @handle if present in either column
                m = re.search(r"@([A-Za-z0-9_.]+)", handle_or_name_raw + " " + real_name)
                if m:
                    handle = "@" + m.group(1)
                platform = first_nonempty(r, 3)
                followers_raw_cell = cell(r, 4)
                tier_label = normalise_tier_label(cell(r, 6))
                if cell(r, 5):
                    notes_parts.append(f"niche: {cell(r, 5)}")
                if tier_label:
                    notes_parts.append(f"editorial-tier: {tier_label}")
                if cell(r, 7):
                    notes_parts.append(f"angle: {cell(r, 7)}")
                if cell(r, 8):
                    notes_parts.append(f"method: {cell(r, 8)}")
                if cell(r, 9):
                    notes_parts.append(f"contact: {cell(r, 9)}")
            else:
                continue

            if not name:
                continue

            followers_num, followers_raw_str = parse_followers(followers_raw_cell)
            tier = tier_from_followers(followers_num)
            handle_key = handle.lower().lstrip("@") if handle else ""
            name_key = slug(name)
            is_dupe = False
            if handle_key and handle_key in seen_handles:
                is_dupe = True
            elif name_key in seen_names:
                # Name match across workbooks counts as a duplicate even if
                # no handle was recorded on one side (e.g. Nolan Perkins).
                is_dupe = True
            if is_dupe:
                skipped_dupes += 1
                continue
            if handle_key:
                seen_handles.add(handle_key)
            seen_names.add(name_key)

            row_id = slug(name)
            if handle:
                row_id += "--" + slug(handle.lstrip("@"))

            # Partner tier: Outreach List entries with editorial Tier 1 get
            # flagship deal (40/35/30); everything else is standard (20%).
            partner_tier = "standard"
            if sheet_name == "Outreach List":
                tlabel = normalise_tier_label(cell(r, 6))
                if tlabel in ("tier-1", "tier-2"):
                    partner_tier = "flagship"

            rows.append({
                "id": row_id,
                "name": name,
                "handle": handle,
                "platform": platform,
                "followers": followers_num if followers_num is not None else "",
                "followers_raw": followers_raw_str,
                "tier": tier,
                "audience_category": category,
                "persona": "",
                "priority": normalise_priority(priority_raw),
                "partner_tier": partner_tier,
                "status": "not-contacted",
                "outreach_channel": "",
                "outreach_date": "",
                "batch_code": "",
                "response": "",
                "notes": "; ".join(notes_parts),
                "source_sheet": sheet_name,
                "source_workbook": workbook_tag,
            })

    return skipped_dupes


def process_seed_csv(
    csv_path: Path,
    rows: list[dict[str, Any]],
    seen_handles: set[str],
    seen_names: set[str],
) -> int:
    """Append rows from a seed CSV (research output). Returns dupes skipped.

    Expected columns: name, handle, platform, followers_raw,
    audience_category, persona, priority, partner_tier, notes.
    """
    skipped_dupes = 0
    workbook_tag = csv_path.stem
    with csv_path.open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            name = (r.get("name") or "").strip()
            if not name:
                continue
            handle = normalise_handle(r.get("handle"))
            handle_key = handle.lower().lstrip("@") if handle else ""
            name_key = slug(name)
            if (handle_key and handle_key in seen_handles) or name_key in seen_names:
                skipped_dupes += 1
                continue
            if handle_key:
                seen_handles.add(handle_key)
            seen_names.add(name_key)

            followers_num, followers_raw_str = parse_followers(r.get("followers_raw"))
            tier = tier_from_followers(followers_num)
            row_id = slug(name)
            if handle:
                row_id += "--" + slug(handle.lstrip("@"))

            rows.append({
                "id": row_id,
                "name": name,
                "handle": handle,
                "platform": (r.get("platform") or "").strip(),
                "followers": followers_num if followers_num is not None else "",
                "followers_raw": followers_raw_str,
                "tier": tier,
                "audience_category": (r.get("audience_category") or "").strip(),
                "persona": (r.get("persona") or "").strip(),
                "priority": (r.get("priority") or "").strip(),
                "partner_tier": (r.get("partner_tier") or "standard").strip(),
                "status": "not-contacted",
                "outreach_channel": "",
                "outreach_date": "",
                "batch_code": "",
                "response": "",
                "notes": (r.get("notes") or "").strip(),
                "source_sheet": "research",
                "source_workbook": workbook_tag,
            })
    return skipped_dupes


def main() -> int:
    paths = [Path(p) for p in sys.argv[1:]] if len(sys.argv) > 1 else DEFAULT_XLSX_PATHS
    missing = [p for p in paths if not p.exists()]
    if missing:
        for p in missing:
            sys.stderr.write(f"Source xlsx not found: {p}\n")
        return 1

    rows: list[dict[str, Any]] = []
    seen_handles: set[str] = set()
    seen_names: set[str] = set()
    skipped_total = 0
    for p in paths:
        print(f"Processing {p}")
        skipped_total += process_workbook(p, rows, seen_handles, seen_names)

    # Seed CSVs (research output, manual additions) — sorted for determinism
    if SEEDS_DIR.exists():
        for seed in sorted(SEEDS_DIR.glob("*.csv")):
            print(f"Processing {seed}")
            skipped_total += process_seed_csv(seed, rows, seen_handles, seen_names)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    by_category: dict[str, int] = {}
    by_tier: dict[str, int] = {}
    by_partner: dict[str, int] = {}
    by_workbook: dict[str, int] = {}
    for r in rows:
        by_category[r["audience_category"]] = by_category.get(r["audience_category"], 0) + 1
        by_tier[r["tier"] or "unknown"] = by_tier.get(r["tier"] or "unknown", 0) + 1
        by_partner[r["partner_tier"]] = by_partner.get(r["partner_tier"], 0) + 1
        by_workbook[r["source_workbook"]] = by_workbook.get(r["source_workbook"], 0) + 1

    print(f"\nWrote {len(rows)} rows to {OUTPUT.relative_to(REPO_ROOT)}")
    print(f"Skipped duplicates across workbooks: {skipped_total}")
    print("\nBy source workbook:")
    for k, v in sorted(by_workbook.items(), key=lambda kv: -kv[1]):
        print(f"  {k:<45} {v}")
    print("\nBy audience category:")
    for k in sorted(by_category, key=lambda k: -by_category[k]):
        print(f"  {k or '(uncategorised)':<22} {by_category[k]}")
    print("\nBy follower tier:")
    for k in ["mega", "macro", "micro", "nano", "unknown"]:
        if k in by_tier:
            print(f"  {k:<8} {by_tier[k]}")
    print("\nBy partner tier (commission deal):")
    for k, v in sorted(by_partner.items(), key=lambda kv: -kv[1]):
        print(f"  {k:<10} {v}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
